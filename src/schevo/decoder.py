import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm

from constants import EXCEL_FORMATS, EXCEL_MAX_ROWS, PATH_CFG, PATH_RES
from core import decode_json


def decode_config(working_dir: str | Path) -> dict[str, dict]:
    """
    Return each file in the working directory that match a stream filename regex into the JSON config file.

    :param working_dir: The directory where search the file to work.
    :type working_dir: str | Path
    :return: A dictionary with file list, encoding and his record configurations for each stream.
    :rtype: dict[str, dict]
    """
    config = decode_json(PATH_CFG / 'schevo.json')
    patterns = {
        stream: re.compile(obj['filename'])
        for stream, obj in config.items()
        if 'filename' in obj
    }

    res = {}
    if isinstance(working_dir, str):
        working_dir = Path(working_dir).resolve()

    for fin in (f for f in working_dir.iterdir() if f.is_file()):
        for stream, regex in patterns.items():
            if regex.match(fin.name):
                res.setdefault(stream, {
                    'streams': [],
                    'encoding': config[stream].get('encoding'),
                    'clean': config[stream].get('clean', False),
                    'record_code': config[stream]['record_code'],
                    'config': config[stream]['records']
                })['streams'].append(fin)
    return res


def decode_record(record: str,
                  config: dict[str, dict] | None) -> dict[str, Any] | None:
    """
    Convert a record row into dictionary by splitting it for each column configured, with the correct data type.

    :param record: The raw record read from the file.
    :type record: str
    :param config: The column configuration read from the JSON config file, already filtered by record code.
    :type config: dict[str, dict] | None
    :return: A dictionary with the original record split in column.
    :rtype: dict[str, Any] | None
    """
    if not config: return None
    res = {}
    for key, field in config.items():
        value = record[field['begin'] - 1 : field['end']].strip()

        match field.get('type'):
            case 'datetime' | 'date' | 'time' if not value.strip('0'):
                value = None
            case 'datetime' if value:
                value = datetime.strptime(value, field.get('format', '%Y%m%d%H%M%S'))
            case 'date' if value:
                value = datetime.strptime(value, field.get('format', '%Y%m%d')).date()
            case 'time' if value:
                value = datetime.strptime(value, field.get('format', '%H%M%S')).time()
            case 'integer' if value:
                value = int(value)
            case 'decimal' if value:
                value = Decimal(value + field.get('format', 'e-2'))
            case _: pass

        res[key] = (value
                    if value
                        and isinstance(value, str)
                    else None)
    return res


def txt2xlsx(streams: dict[str, dict]) -> None:
    """
    Convert a file with format fixed-length into Excel file, one or more sheet for each record code.

    :param streams: The result of decode_config() function with the input file and stream configuration.
    :type streams: dict[str, dict]
    """
    for stream, config in streams.items():
        for fin in config['streams']:
            total_rows = sum(1 for _ in open(fin, encoding=config.get('encoding')))

            with open(fin, encoding=config.get('encoding')) as source_in:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

                indexes = {}
                for row in tqdm(source_in,
                                total=total_rows,
                                bar_format='{n_fmt}/{total_fmt} |{bar}| {percentage:3.0f}%'):
                    if config['record_code']:
                        record_code = row[config['record_code'][0] - 1 : config['record_code'][1]]
                    else: record_code = stream
                    record = decode_record(row, config['config'].get(record_code))
                    if not record: continue

                    if record_code not in indexes:
                        indexes[record_code] = {'row': 2, 'index': 1}
                    ws_name = record_code
                    if indexes[record_code]['index'] > 1:
                        ws_name = f'{record_code}_{indexes[record_code]['index']:02d}'

                    if record_code in wb.sheetnames:
                        ws = wb[ws_name]
                    else:
                        ws = wb.create_sheet(ws_name)
                        for col_num, col in enumerate([
                            k.upper().replace('_', ' ')
                            for k in config['config'][record_code].keys()
                        ], start=1):
                            ws.cell(row=1, column=col_num).value = col
                            ws.cell(row=1, column=col_num).font = Font(name='Aptos Narrow', bold=True)
                            ws.cell(row=1, column=col_num).number_format = '@'

                    row_num = indexes[record_code]['row']
                    if row_num > EXCEL_MAX_ROWS:
                        if indexes[record_code]['index'] == 1:
                            wb[ws_name].title = f'{record_code}_01'
                        indexes[record_code]['index'] += 1
                        ws_name = f'{record_code}_{indexes[record_code]['index']:02d}'
                        ws = wb.create_sheet(ws_name)

                        for col_num, col in enumerate([
                            k.upper().replace('_', ' ')
                            for k in config['config'][record_code].keys()
                        ], start=1):
                            ws.cell(row=1, column=col_num).value = col
                            ws.cell(row=1, column=col_num).font = Font(name='Aptos Narrow', bold=True)
                            ws.cell(row=1, column=col_num).number_format = '@'

                        row_num = 2

                    for col_num, cell in enumerate(record.values(), start=1):
                        ws.cell(row=row_num, column=col_num).value = cell
                        ws.cell(row=row_num, column=col_num).font = Font(name='Aptos Narrow')
                        ws.cell(row=row_num, column=col_num).number_format = EXCEL_FORMATS.get(type(cell))
                    indexes[record_code]['row'] = row_num + 1

                for ws in wb.worksheets:
                    ws.auto_filter.ref = ws.dimensions
                    ws.freeze_panes = 'A2'
                wb.save(fin.with_suffix('.xlsx'))

if __name__ == '__main__':
    job_begin = datetime.now()
    txt2xlsx(decode_config(PATH_RES))
    print(datetime.now() - job_begin)
